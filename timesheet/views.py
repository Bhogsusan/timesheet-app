# timesheet/views.py (updated with Company CRUD views)
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.views.generic import ListView, CreateView, DetailView, DeleteView, UpdateView
from django.urls import reverse_lazy
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import date, timedelta
import calendar

from .models import Company, Timesheet, TimesheetEntry, ExtraHours
from .forms import TimesheetForm, CompanySelectForm, ExtraHoursForm, ExtraHoursFormSet, CompanyForm

# ==================== COMPANY VIEWS (Website Management) ====================

class CompanyListView(ListView):
    model = Company
    template_name = 'timesheet/company_list.html'
    context_object_name = 'companies'
    ordering = ['name']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_count'] = Company.objects.filter(is_active=True).count()
        context['inactive_count'] = Company.objects.filter(is_active=False).count()
        return context


class CompanyCreateView(CreateView):
    model = Company
    form_class = CompanyForm
    template_name = 'timesheet/company_form.html'
    success_url = reverse_lazy('company_list')
    
    def form_valid(self, form):
        messages.success(self.request, f'Company "{form.instance.name}" created successfully!')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Add New Company'
        context['action'] = 'Create'
        return context


class CompanyDetailView(DetailView):
    model = Company
    template_name = 'timesheet/company_detail.html'
    context_object_name = 'company'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.object
        
        # Calculate weekly totals
        if company.pattern_type == 'weekly':
            context['weekly_total'] = (
                company.mon_hours + company.tue_hours + company.wed_hours +
                company.thu_hours + company.fri_hours + company.sat_hours + company.sun_hours
            )
        else:
            context['week_a_total'] = (
                company.mon_hours_week_a + company.tue_hours_week_a + company.wed_hours_week_a +
                company.thu_hours_week_a + company.fri_hours_week_a + company.sat_hours_week_a + company.sun_hours_week_a
            )
            context['week_b_total'] = (
                company.mon_hours_week_b + company.tue_hours_week_b + company.wed_hours_week_b +
                company.thu_hours_week_b + company.fri_hours_week_b + company.sat_hours_week_b + company.sun_hours_week_b
            )
        
        # Get timesheets using this company
        context['timesheet_count'] = TimesheetEntry.objects.filter(company=company).count()
        
        return context


class CompanyUpdateView(UpdateView):
    model = Company
    form_class = CompanyForm
    template_name = 'timesheet/company_form.html'
    success_url = reverse_lazy('company_list')
    
    def form_valid(self, form):
        messages.success(self.request, f'Company "{form.instance.name}" updated successfully!')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Edit {self.object.name}'
        context['action'] = 'Update'
        context['company'] = self.object
        return context


class CompanyDeleteView(DeleteView):
    model = Company
    template_name = 'timesheet/company_confirm_delete.html'
    success_url = reverse_lazy('company_list')
    
    def delete(self, request, *args, **kwargs):
        company = self.get_object()
        messages.success(request, f'Company "{company.name}" deleted successfully!')
        return super().delete(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Check if company is used in any timesheets
        context['timesheet_count'] = TimesheetEntry.objects.filter(company=self.object).count()
        return context


# ==================== TIMESHEET VIEWS ====================

class TimesheetListView(ListView):
    model = Timesheet
    template_name = 'timesheet/timesheet_list.html'
    context_object_name = 'timesheets'
    ordering = ['-year', '-month']

def create_timesheet(request):
    if request.method == 'POST':
        form = TimesheetForm(request.POST)
        company_ids = request.POST.getlist('companies')
        
        if form.is_valid() and company_ids:
            timesheet = form.save()
            
            # Create entries for selected companies
            for company_id in company_ids:
                company = Company.objects.get(id=company_id)
                TimesheetEntry.objects.create(timesheet=timesheet, company=company)
            
            # Handle extra hours - only process if data is provided
            extra_hours_dates = request.POST.getlist('extra_hours_date')
            extra_hours_companies = request.POST.getlist('extra_hours_company')
            extra_hours_amounts = request.POST.getlist('extra_hours_amount')
            extra_hours_descs = request.POST.getlist('extra_hours_description')
            
            for i in range(len(extra_hours_dates)):
                date_val = extra_hours_dates[i] if i < len(extra_hours_dates) else ''
                hours_val = extra_hours_amounts[i] if i < len(extra_hours_amounts) else ''
                desc_val = extra_hours_descs[i] if i < len(extra_hours_descs) else ''
                company_val = extra_hours_companies[i] if i < len(extra_hours_companies) else ''
                
                # Only create if we have at least date and hours
                if date_val and hours_val:
                    try:
                        hours_decimal = Decimal(str(hours_val))
                        if hours_decimal > 0:  # Only save if hours > 0
                            ExtraHours.objects.create(
                                timesheet=timesheet,
                                company_id=int(company_val) if company_val else None,
                                date=date_val,
                                hours=hours_decimal,
                                description=desc_val
                            )
                    except (ValueError, InvalidOperation):
                        pass  # Skip invalid data
            
            messages.success(request, 'Timesheet created successfully!')
            return redirect('timesheet_detail', pk=timesheet.pk)
    else:
        form = TimesheetForm()
    
    companies = Company.objects.filter(is_active=True)
    return render(request, 'timesheet/timesheet_form.html', {
        'form': form,
        'companies': companies,
    })

def timesheet_detail(request, pk):
    timesheet = get_object_or_404(Timesheet, pk=pk)
    entries = timesheet.entries.select_related('company').all()
    extra_hours = timesheet.extra_hours.all()
    
    # Generate calendar data
    year = timesheet.year
    month = timesheet.month
    _, days_in_month = calendar.monthrange(year, month)
    
    calendar_data = []
    for day in range(1, days_in_month + 1):
        current_date = date(year, month, day)
        weekday = current_date.weekday()
        day_name = calendar.day_abbr[weekday]
        
        row = {
            'date': day,
            'day_name': day_name,
            'companies': []
        }
        
        for entry in entries:
            hours = entry.get_daily_hours(day)
            if hours > 0:
                row['companies'].append({
                    'name': entry.company.name,
                    'hours': hours
                })
        
        calendar_data.append(row)
    
    # Calculate totals
    company_totals = {}
    grand_total = 0
    
    for entry in entries:
        total = entry.get_total_hours()
        company_totals[entry.company.name] = total
        grand_total += total
    
    extra_hours_total = sum(float(eh.hours) for eh in extra_hours)
    grand_total += extra_hours_total
    
    context = {
        'timesheet': timesheet,
        'entries': entries,
        'extra_hours': extra_hours,
        'calendar_data': calendar_data,
        'company_totals': company_totals,
        'grand_total': grand_total,
        'days_in_month': days_in_month,
    }
    
    return render(request, 'timesheet/timesheet_detail.html', context)


    entries = timesheet.entries.select_related('company').all()
    extra_hours = timesheet.extra_hours.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"
    
    # Styles
    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=12)
    total_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    white_font = Font(bold=True, color="FFFFFF")
    
    # Title
    ws['A1'] = f"Cleaner: {timesheet.cleaner_name}"
    ws['A1'].font = title_font
    ws['A2'] = f"Month: {calendar.month_name[timesheet.month]} {timesheet.year}"
    ws['A2'].font = title_font
    
    # Headers
    headers = ['Date', 'Day'] + [entry.company.name for entry in entries]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Data rows
    year = timesheet.year
    month = timesheet.month
    _, days_in_month = calendar.monthrange(year, month)
    
    row_num = 5
    for day in range(1, days_in_month + 1):
        current_date = date(year, month, day)
        weekday = current_date.weekday()
        day_name = calendar.day_abbr[weekday]
        
        ws.cell(row=row_num, column=1, value=day).border = thin_border
        ws.cell(row=row_num, column=2, value=day_name).border = thin_border
        
        col_num = 3
        for entry in entries:
            hours = entry.get_daily_hours(day)
            cell = ws.cell(row=row_num, column=col_num, value=hours if hours > 0 else 0)
            cell.border = thin_border
            cell.alignment = center_align
            col_num += 1
        
        row_num += 1
    
    # Total row
    total_row = row_num
    ws.cell(row=total_row, column=1, value="TOTAL").font = total_font
    ws.cell(row=total_row, column=2, value="").border = thin_border
    
    col_num = 3
    grand_total = 0
    for entry in entries:
        total = entry.get_total_hours()
        grand_total += total
        cell = ws.cell(row=total_row, column=col_num, value=total)
        cell.font = total_font
        cell.border = thin_border
        cell.alignment = center_align
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        col_num += 1
    
    # Grand Total row
    grand_total_row = total_row + 1
    ws.merge_cells(start_row=grand_total_row, start_column=1, end_row=grand_total_row, end_column=2)
    grand_total_cell = ws.cell(row=grand_total_row, column=1, value="GRAND TOTAL")
    grand_total_cell.font = white_font
    grand_total_cell.fill = blue_fill
    grand_total_cell.alignment = center_align
    grand_total_cell.border = thin_border
    
    ws.merge_cells(start_row=grand_total_row, start_column=3, end_row=grand_total_row, end_column=col_num-1)
    grand_value_cell = ws.cell(row=grand_total_row, column=3, value=grand_total)
    grand_value_cell.font = white_font
    grand_value_cell.fill = blue_fill
    grand_value_cell.alignment = center_align
    grand_value_cell.border = thin_border
    
    # Extra Hours section
    extra_row = grand_total_row + 2
    if extra_hours.exists():
        ws.cell(row=extra_row, column=1, value="Extra Hours").font = title_font
        extra_row += 1
        
        for eh in extra_hours:
            ws.cell(row=extra_row, column=1, value=eh.date.day)
            ws.cell(row=extra_row, column=2, value=calendar.day_abbr[eh.date.weekday()])
            ws.cell(row=extra_row, column=3, value=float(eh.hours))
            ws.cell(row=extra_row, column=4, value=eh.description)
            extra_row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 10
    for i in range(3, col_num):
        ws.column_dimensions[chr(64+i)].width = 15
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{timesheet.cleaner_name}_{calendar.month_name[timesheet.month]}_{timesheet.year}.xlsx"'
    
    wb.save(response)
    return response


def generate_excel(request, pk):
    from decimal import Decimal
    timesheet = get_object_or_404(Timesheet, pk=pk)
    entries = timesheet.entries.select_related('company').all()
    extra_hours = timesheet.extra_hours.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"
    
    # Styles
    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=12)
    total_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    white_font = Font(bold=True, color="FFFFFF")
    
    # Title
    ws['A1'] = f"Cleaner: {timesheet.cleaner_name}"
    ws['A1'].font = title_font
    ws['A2'] = f"Month: {timesheet.get_month_name()} {timesheet.year}"  # FIXED
    ws['A2'].font = title_font
    
    # Headers
    headers = ['Date', 'Day'] + [entry.company.name for entry in entries]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Data rows
    year = timesheet.year
    month = timesheet.month
    _, days_in_month = calendar.monthrange(year, month)
    
    row_num = 5
    for day in range(1, days_in_month + 1):
        current_date = date(year, month, day)
        weekday = current_date.weekday()
        day_name = calendar.day_abbr[weekday]
        
        ws.cell(row=row_num, column=1, value=day).border = thin_border
        ws.cell(row=row_num, column=2, value=day_name).border = thin_border
        
        col_num = 3
        for entry in entries:
            hours = entry.get_daily_hours(day)
            # FIXED: Convert Decimal to float for Excel, properly formatted
            hours_val = float(hours) if hours else None
            cell = ws.cell(row=row_num, column=col_num, value=hours_val)
            cell.border = thin_border
            cell.alignment = center_align
            cell.number_format = '0.00'  # FIXED: Ensure 2 decimal places in Excel
            col_num += 1
        
        row_num += 1
    
    # Total row
    total_row = row_num
    ws.cell(row=total_row, column=1, value="TOTAL").font = total_font
    ws.cell(row=total_row, column=2, value="").border = thin_border
    
    col_num = 3
    grand_total = Decimal('0.00')
    for entry in entries:
        total = entry.get_total_hours()
        grand_total += total
        cell = ws.cell(row=total_row, column=col_num, value=float(total))
        cell.font = total_font
        cell.border = thin_border
        cell.alignment = center_align
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        cell.number_format = '0.00'  # FIXED
        col_num += 1
    
    # Grand Total row
    grand_total_row = total_row + 1
    ws.merge_cells(start_row=grand_total_row, start_column=1, end_row=grand_total_row, end_column=2)
    grand_total_cell = ws.cell(row=grand_total_row, column=1, value="GRAND TOTAL")
    grand_total_cell.font = white_font
    grand_total_cell.fill = blue_fill
    grand_total_cell.alignment = center_align
    grand_total_cell.border = thin_border
    
    ws.merge_cells(start_row=grand_total_row, start_column=3, end_row=grand_total_row, end_column=col_num-1)
    grand_value_cell = ws.cell(row=grand_total_row, column=3, value=float(grand_total))
    grand_value_cell.font = white_font
    grand_value_cell.fill = blue_fill
    grand_value_cell.alignment = center_align
    grand_value_cell.border = thin_border
    grand_value_cell.number_format = '0.00'  # FIXED
    
    # Extra Hours section
    extra_row = grand_total_row + 2
    if extra_hours.exists():
        ws.cell(row=extra_row, column=1, value="Extra Hours").font = title_font
        extra_row += 1
        
        for eh in extra_hours:
            ws.cell(row=extra_row, column=1, value=eh.date.day if eh.date else '')
            ws.cell(row=extra_row, column=2, value=calendar.day_abbr[eh.date.weekday()] if eh.date else '')
            cell = ws.cell(row=extra_row, column=3, value=float(eh.hours))
            cell.number_format = '0.00'  # FIXED
            ws.cell(row=extra_row, column=4, value=eh.description)
            extra_row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 10
    for i in range(3, col_num):
        ws.column_dimensions[chr(64+i)].width = 15
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{timesheet.cleaner_name}_{timesheet.get_month_name()}_{timesheet.year}.xlsx"'  # FIXED
    
    wb.save(response)
    return response


def delete_timesheet(request, pk):
    timesheet = get_object_or_404(Timesheet, pk=pk)
    if request.method == 'POST':
        timesheet.delete()
        messages.success(request, 'Timesheet deleted successfully!')
        return redirect('timesheet_list')
    return render(request, 'timesheet/timesheet_confirm_delete.html', {'timesheet': timesheet})

def get_company_preview(request):
    company_id = request.GET.get('company_id')
    if company_id:
        company = get_object_or_404(Company, id=company_id)
        data = {
            'pattern_type': company.pattern_type,
            'weekly': {
                'mon': float(company.mon_hours),
                'tue': float(company.tue_hours),
                'wed': float(company.wed_hours),
                'thu': float(company.thu_hours),
                'fri': float(company.fri_hours),
                'sat': float(company.sat_hours),
                'sun': float(company.sun_hours),
            },
            'biweekly_a': {
                'mon': float(company.mon_hours_week_a),
                'tue': float(company.tue_hours_week_a),
                'wed': float(company.wed_hours_week_a),
                'thu': float(company.thu_hours_week_a),
                'fri': float(company.fri_hours_week_a),
                'sat': float(company.sat_hours_week_a),
                'sun': float(company.sun_hours_week_a),
            },
            'biweekly_b': {
                'mon': float(company.mon_hours_week_b),
                'tue': float(company.tue_hours_week_b),
                'wed': float(company.wed_hours_week_b),
                'thu': float(company.thu_hours_week_b),
                'fri': float(company.fri_hours_week_b),
                'sat': float(company.sat_hours_week_b),
                'sun': float(company.sun_hours_week_b),
            },
        }
        return JsonResponse(data)
    return JsonResponse({'error': 'No company ID provided'}, status=400)